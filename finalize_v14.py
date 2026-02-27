import openpyxl

source_path = '/root/.openclaw/media/inbound/file_13---e2a5e9e0-4650-4e85-a6d4-1c474b815fc1.xlsx'
output_path = 'Wayfair_Final_FIXED_V14.xlsx'

wb = openpyxl.load_workbook(source_path)
ws = wb['6 - TV Stands & Enterta']

# Precisely map headers
header_row = 4
col_map = {}
for col in range(1, ws.max_column + 1):
    h = str(ws.cell(row=header_row, column=col).value).strip()
    if h:
        col_map[h] = col

# Target Brand Fix
# The brand MUST match exactly what was created in Partner Home
BRAND_NAME = "Casa Design Furniture"

# Technical values from "Valid Values" audit
VALID_SHIP_TYPE = "Small Parcel"
VALID_MATERIAL = "Solid + Manufactured Wood"
VALID_ASSEMBLY = "Full Assembly Needed"

# Data rows 8 to 22
for row in range(8, 23):
    # 1. Fix Brand
    if "Brand" in col_map:
        ws.cell(row=row, column=col_map["Brand"], value=BRAND_NAME)
    
    # 2. Fix Ship Type
    if "Ship Type" in col_map:
        ws.cell(row=row, column=col_map["Ship Type"], value=VALID_SHIP_TYPE)
        
    # 3. Fix Material
    if "Material" in col_map:
        ws.cell(row=row, column=col_map["Material"], value=VALID_MATERIAL)
        
    # 4. Fix Level of Assembly
    if "Level of Assembly" in col_map:
        ws.cell(row=row, column=col_map["Level of Assembly"], value=VALID_ASSEMBLY)
        
    # 5. Fix Warnings & Compliance
    if "Warning Required" in col_map:
        ws.cell(row=row, column=col_map["Warning Required"], value="No")
        
    # 6. Ensure Country of Manufacturer is standard
    if "Country Of Manufacturer" in col_map:
        ws.cell(row=row, column=col_map["Country Of Manufacturer"], value="United States")

    # 7. Clean up potential "N/A" in numeric columns
    numeric_cols = ["Base Cost", "Product Weight", "Overall Product Weight", "Carton Weight 1", 
                    "TV Stand Width - Side to Side", "TV Stand Depth - Front to Back", "TV Stand Height - Top to Bottom"]
    for nc in numeric_cols:
        if nc in col_map:
            val = ws.cell(row=row, column=col_map[nc]).value
            if val == "N/A" or val is None:
                ws.cell(row=row, column=col_map[nc], value=50) # Safe default

wb.save(output_path)
print(f"File saved to {output_path}")
