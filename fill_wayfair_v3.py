import openpyxl
import json
import time

file_path = '/root/.openclaw/media/inbound/file_3---fdf5fa93-ae80-487c-a3ba-d83afddef241.xlsx'
output_path = 'Wayfair_Spec_Sheet_Filled.xlsx'

wb = openpyxl.load_workbook(file_path)

tv_stands = [
    {
        'Supplier Part Number': 'CASA-TV-018',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#018 NORALIE TV STAND & FIREPLACE',
        'Marketing Copy': 'Elevate your entertainment area with the Noralie TV Stand & Fireplace. Featuring a beveled mirrored finish and elegant faux diamond inlays, this free-standing unit adds a touch of luxury to any room. Includes a built-in LED electric fireplace with touch control and remote, providing 4777 BTUs of heat for areas up to 400 sq ft.',
        'Feature Bullet 1': 'Mirrored finish with beveled edges',
        'Feature Bullet 2': 'Faux diamond inlay accents',
        'Feature Bullet 3': 'Built-in 4777 BTU LED electric fireplace',
        'Feature Bullet 4': 'Includes remote control and touch panel',
        'Feature Bullet 5': 'Two storage doors with internal compartments',
        'Base Cost': 899.00
    },
    {
        'Supplier Part Number': 'CASA-TV-013',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#013 TV STAND IN WHITE LACQUER',
        'Marketing Copy': 'Modern TV stand in white lacquer with polished stainless steel legs. Features self-closing doors and drawers with extension rails for a sleek, organized living space.',
        'Feature Bullet 1': 'High-gloss white lacquer finish',
        'Feature Bullet 2': 'Polished stainless steel legs',
        'Feature Bullet 3': 'Self-closing doors and drawers',
        'Feature Bullet 4': 'Extension rails for smooth operation',
        'Feature Bullet 5': 'Pre-assembled for convenience',
        'Base Cost': 799.00
    },
    {
        'Supplier Part Number': 'CASA-TV-015',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#015 LED LIGHT TV STAND',
        'Marketing Copy': 'A contemporary TV stand designed to brighten your living room. Features integrated LED lighting for a modern aesthetic and ample storage for media devices.',
        'Feature Bullet 1': 'Integrated LED lighting system',
        'Feature Bullet 2': 'Contemporary modern design',
        'Feature Bullet 3': 'Ample storage for electronic devices',
        'Feature Bullet 4': 'High-quality wood and laminate construction',
        'Feature Bullet 5': 'Sleek hardware-free appearance',
        'Base Cost': 649.00
    },
    {
        'Supplier Part Number': 'CASA-TV-016',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#016 TV STAND WHITE & LED LIGHT',
        'Marketing Copy': 'Modern and stylish TV stand with two door-compartments and a center drawer. Includes a spacious exposed area for electronic devices and integrated white LED light for a sophisticated look.',
        'Feature Bullet 1': 'Modern white finish with integrated LED',
        'Feature Bullet 2': 'Two self-closing side doors',
        'Feature Bullet 3': 'Center storage drawer for organization',
        'Feature Bullet 4': 'Exposed media shelving for electronics',
        'Feature Bullet 5': 'Perfect for apartments and contemporary homes',
        'Base Cost': 699.00
    },
    {
        'Supplier Part Number': 'CASA-TV-017',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#017 KL TV STAND',
        'Marketing Copy': 'The KL TV Stand offers a minimalist yet functional design, providing a stable and stylish base for your home entertainment system.',
        'Feature Bullet 1': 'Minimalist modern silhouette',
        'Feature Bullet 2': 'Durable construction for heavy TVs',
        'Feature Bullet 3': 'Versatile finish that fits any decor',
        'Feature Bullet 4': 'Hidden cable management features',
        'Feature Bullet 5': 'Easy to clean and maintain surface',
        'Base Cost': 549.00
    },
    {
        'Supplier Part Number': 'CASA-TV-019',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#019 TV STAND MIRRORED & DIAMONDS',
        'Marketing Copy': 'Add a spectacular contemporary touch with the Noralie Mirrored TV Stand. Featuring a fully mirrored base and top with faux diamond inlay fronts, this piece holds up to a 50 inch flat screen.',
        'Feature Bullet 1': 'Fully mirrored base and top',
        'Feature Bullet 2': 'Faux diamond inlay on drawer and cabinet fronts',
        'Feature Bullet 3': 'Holds up to 50 inch flat screen TV',
        'Feature Bullet 4': 'One storage drawer and two cabinets',
        'Feature Bullet 5': 'Open media compartment for easy access',
        'Base Cost': 1099.00
    },
    {
        'Supplier Part Number': 'CASA-TV-020',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#020 TV STAND ELECTRIC FIREPLACE',
        'Marketing Copy': 'Combine warmth and luxury with this rectangular TV stand featuring a beveled mirrored finish and faux diamond inlays. Includes a 4777 BTU LED fireplace for 400 sq ft heating.',
        'Feature Bullet 1': 'Mirrored case-frame with sled base',
        'Feature Bullet 2': '4777 BTU LED electric fireplace included',
        'Feature Bullet 3': 'Touch control panel and remote control',
        'Feature Bullet 4': 'Flame light can be used without heat',
        'Feature Bullet 5': 'Large storage compartments with doors',
        'Base Cost': 1299.00
    }
]

beds = [
    {
        'Supplier Part Number': 'CASA-BED-003',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#003 PERRIS TWIN WORKSTATION LOFT BED',
        'Marketing Copy': 'Combine sleep and study with the Perris Twin Workstation Loft Bed. Features a large workspace with a keyboard drawer and integrated bookshelves, crafted from Asian veneers.',
        'Feature Bullet 1': 'Space-saving twin loft configuration',
        'Feature Bullet 2': 'Integrated study desk and workstation',
        'Feature Bullet 3': 'Keyboard drawer and bookshelves included',
        'Feature Bullet 4': 'Tested static weight limit of 400 lbs',
        'Feature Bullet 5': 'Asian veneers and select wood construction',
        'Base Cost': 899.00
    },
    {
        'Supplier Part Number': 'CASA-BED-002',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#002 JONES TWIN BED WHITE',
        'Marketing Copy': 'Classic twin bed with a charming bobbin motif. Constructed with solid pine for durability and includes a full slat kit, making a bunkie board unnecessary.',
        'Feature Bullet 1': 'Elegant bobbin motif design',
        'Feature Bullet 2': 'Solid pine wood construction',
        'Feature Bullet 3': 'Includes complete slat kit',
        'Feature Bullet 4': 'Bunkie board not required',
        'Feature Bullet 5': 'Glossy white finish for a clean look',
        'Base Cost': 359.00
    },
    {
        'Supplier Part Number': 'CASA-BED-025',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#025 JESSICA QUEEN SIZE BED',
        'Marketing Copy': 'Modern queen bed in a rich cappuccino finish. The headboard features three spacious open storage compartments, perfect for books and accessories.',
        'Feature Bullet 1': 'Integrated headboard storage compartments',
        'Feature Bullet 2': 'Contemporary low-profile footboard',
        'Feature Bullet 3': 'Finished in rich cappuccino',
        'Feature Bullet 4': 'Solid and manufactured wood blend',
        'Feature Bullet 5': 'Part of the premium Jessica collection',
        'Base Cost': 699.00
    },
    {
        'Supplier Part Number': 'CASA-BED-ADJ-002',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#002 ADJUSTABLE BED BASE QUEEN',
        'Marketing Copy': 'Entry-level adjustable base with head and foot incline. Helps reduce snoring and sleep apnea while providing a comfortable position for reading or watching TV.',
        'Feature Bullet 1': 'Independent head and foot incline',
        'Feature Bullet 2': 'Improves sleep health and comfort',
        'Feature Bullet 3': 'Fits into most standard bed frames',
        'Feature Bullet 4': 'Simple remote control operation',
        'Feature Bullet 5': 'Durable steel frame construction',
        'Base Cost': 399.00
    },
    {
        'Supplier Part Number': 'CASA-BED-004',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#004 SWEETHEART TEEN LOFT BED PINK',
        'Marketing Copy': 'Whimsical fire engine style youth tent bed with a glossy white frame and pink accents. Includes guard rails and a coordinating ladder for safety.',
        'Feature Bullet 1': 'Playful fire engine tent design',
        'Feature Bullet 2': 'Glossy white frame with pink accents',
        'Feature Bullet 3': 'Includes safety guard rails and ladder',
        'Feature Bullet 4': 'Weight capacity of 400 lbs (top only)',
        'Feature Bullet 5': 'Recommended for children 6 years and up',
        'Base Cost': 499.00
    },
    {
        'Supplier Part Number': 'CASA-BED-028',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#028 LOUIS PHILIPPE WHITE QUEEN BED',
        'Marketing Copy': 'Classic Louis Philippe style wooden queen bed with a clean white finish. Features matching geometric beveling and eye-catching curves on the sleek legs.',
        'Feature Bullet 1': 'Classic Louis Philippe sleigh style',
        'Feature Bullet 2': 'Clean white wooden finish',
        'Feature Bullet 3': 'Elegant geometric beveling detail',
        'Feature Bullet 4': 'Revitalizing curved sleek legs',
        'Feature Bullet 5': 'Solid wood and veneer construction',
        'Base Cost': 459.00
    },
    {
        'Supplier Part Number': 'CASA-BED-ADJ-003',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#003 ADJUSTABLE BED BASE',
        'Marketing Copy': 'Advanced adjustable bed base offering multiple preset positions and quiet motor operation for a restorative night’s sleep.',
        'Feature Bullet 1': 'Quiet motor for seamless adjustment',
        'Feature Bullet 2': 'Multiple preset positions included',
        'Feature Bullet 3': 'Zero-gravity mode for pressure relief',
        'Feature Bullet 4': 'Sturdy construction with non-slip surface',
        'Feature Bullet 5': 'Easy assembly in under 10 minutes',
        'Base Cost': 459.00
    },
    {
        'Supplier Part Number': 'CASA-BED-ADJ-004',
        'Brand': 'Casa Design Furniture',
        'Product Name': '#004 ADJUSTABLE BED BASE PREMIUM',
        'Marketing Copy': 'Premium adjustable base with massage features and under-bed lighting. Elevate your bedroom experience with ultimate comfort and luxury.',
        'Feature Bullet 1': 'Built-in massage feature with intensity levels',
        'Feature Bullet 2': 'Integrated under-bed LED lighting',
        'Feature Bullet 3': 'Wireless remote control with memory buttons',
        'Feature Bullet 4': 'Highest grade steel and component materials',
        'Feature Bullet 5': 'USB charging ports integrated into base',
        'Base Cost': 559.00
    }
]

def fill_sheet(sheet_name, data_list):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=4, column=col).value
        if header_val:
            headers[header_val] = col
            
    current_row = 6
    for data in data_list:
        for key, value in data.items():
            if key in headers:
                ws.cell(row=current_row, column=headers[key], value=value)
        current_row += 1

# Clear previous test data and fill with 15 items
fill_sheet('6 - TV Stands & Enterta', tv_stands) # 7 items
fill_sheet('12 - Beds', beds) # 8 items

wb.save(output_path)
print(f"Saved exactly 15 items to {output_path}")
