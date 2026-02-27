import openpyxl
import json

file_path = '/root/.openclaw/media/inbound/file_3---fdf5fa93-ae80-487c-a3ba-d83afddef241.xlsx'
output_path = 'Wayfair_Spec_Sheet_Filled.xlsx'

wb = openpyxl.load_workbook(file_path)

# Data for TV Stand
tv_stand = {
    'Supplier Part Number': 'CASA-TV-018',
    'Brand': 'Casa Design Furniture',
    'Product Name': '#018 NORALIE TV STAND & FIREPLACE',
    'Marketing Copy': 'Elevate your entertainment area with the Noralie TV Stand & Fireplace. Featuring a beveled mirrored finish and elegant faux diamond inlays, this free-standing unit adds a touch of luxury to any room. Includes a built-in LED electric fireplace with touch control and remote, providing 4777 BTUs of heat for areas up to 400 sq ft.',
    'Feature Bullet 1': 'Mirrored finish with beveled edges',
    'Feature Bullet 2': 'Faux diamond inlay accents',
    'Feature Bullet 3': 'Built-in 4777 BTU LED electric fireplace',
    'Feature Bullet 4': 'Includes remote control and touch panel',
    'Feature Bullet 5': 'Two storage doors with internal compartments',
    'Base Cost': 899.00
}

# Data for Bed
bed = {
    'Supplier Part Number': 'CASA-BED-003',
    'Brand': 'Casa Design Furniture',
    'Product Name': '#003 PERRIS TWIN WORKSTATION LOFT BED',
    'Marketing Copy': 'Perfect for space-saving needs, the Perris Twin Workstation Loft Bed combines sleep and study in one elegant design. Crafted from select woods and Asian veneers, it features a large workspace with a dedicated keyboard drawer and integrated bookshelves.',
    'Feature Bullet 1': 'Space-saving twin loft bed design',
    'Feature Bullet 2': 'Integrated study area with large workspace',
    'Feature Bullet 3': 'Keyboard drawer and storage bookshelves included',
    'Feature Bullet 4': 'Max weight capacity of 400 lbs',
    'Feature Bullet 5': 'Crafted from Asian veneers and select woods',
    'Base Cost': 459.00,
    'Overall Width': 41.75,
    'Overall Depth': 80.0,
    'Overall Height': 74.0
}

def fill_sheet(sheet_name, data):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    
    # Headers are in row 4
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=4, column=col).value
        if header_val:
            headers[header_val] = col
            
    # Row 6 is where data starts
    target_row = 6
    while ws.cell(row=target_row, column=2).value is not None:
        target_row += 1
        
    for key, value in data.items():
        if key in headers:
            ws.cell(row=target_row, column=headers[key], value=value)

# Fill sheets
fill_sheet('6 - TV Stands & Enterta', tv_stand)
fill_sheet('12 - Beds', bed)

wb.save(output_path)
print(f"Saved filled sheet to {output_path}")
