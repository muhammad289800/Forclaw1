import openpyxl
import json

file_path = '/root/.openclaw/media/inbound/file_3---fdf5fa93-ae80-487c-a3ba-d83afddef241.xlsx'
output_path = 'Wayfair_15_Products_Final.xlsx'

wb = openpyxl.load_workbook(file_path)

products = [
    # TV STANDS (Category 6)
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-018', 'Brand': 'Casa Design Furniture', 'Product Name': '#018 NORALIE TV STAND & FIREPLACE', 'Marketing Copy': 'Mirrored finish with beveled edges and faux diamond inlay. Includes LED fireplace.', 'Base Cost': 899}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-013', 'Brand': 'Casa Design Furniture', 'Product Name': '#013 TV STAND IN WHITE LACQUER', 'Marketing Copy': 'Modern white lacquer finish with stainless steel legs. Self-closing doors.', 'Base Cost': 750}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-015', 'Brand': 'Casa Design Furniture', 'Product Name': '#015 LED LIGHT TV STAND', 'Marketing Copy': 'Contemporary design with integrated LED lighting for a modern home aesthetic.', 'Base Cost': 680}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-016', 'Brand': 'Casa Design Furniture', 'Product Name': '#016 TV STAND WHITE & LED LIGHT', 'Marketing Copy': 'Stylish white finish with centralized drawer and spacious storage compartments.', 'Base Cost': 720}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-017', 'Brand': 'Casa Design Furniture', 'Product Name': '#017 KL TV STAND', 'Marketing Copy': 'Elegant KL series TV stand with premium finish and durable construction.', 'Base Cost': 810}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-019', 'Brand': 'Casa Design Furniture', 'Product Name': '#019 TV STAND MIRRORED DIAMONDS', 'Marketing Copy': 'Luxurious mirrored surfaces with sparkling diamond-like accents.', 'Base Cost': 950}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-020', 'Brand': 'Casa Design Furniture', 'Product Name': '#020 TV STAND ELECTRIC FIREPLACE', 'Marketing Copy': 'Functional entertainment center featuring a cozy electric fireplace unit.', 'Base Cost': 1100}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-023', 'Brand': 'Casa Design Furniture', 'Product Name': '#023 AMRITA TV STAND', 'Marketing Copy': 'Modern minimalist Amrita series stand with ample storage space.', 'Base Cost': 590}},
    {'sheet': '6 - TV Stands & Enterta', 'data': {'Supplier Part Number': 'CASA-TV-025', 'Brand': 'Casa Design Furniture', 'Product Name': '#025 AFONSO TV STAND FIREPLACE', 'Marketing Copy': 'Premium Afonso series stand combining heat and style for your living room.', 'Base Cost': 1250}},
    # BEDS (Category 12)
    {'sheet': '12 - Beds', 'data': {'Supplier Part Number': 'CASA-BED-003', 'Brand': 'Casa Design Furniture', 'Product Name': '#003 PERRIS TWIN LOFT BED', 'Marketing Copy': 'Space-saving loft bed with integrated workstation and bookshelves.', 'Base Cost': 459}},
    {'sheet': '12 - Beds', 'data': {'Supplier Part Number': 'CASA-BED-002', 'Brand': 'Casa Design Furniture', 'Product Name': '#002 JONES TWIN BED WHITE', 'Marketing Copy': 'Clean white twin bed perfect for children\'s rooms or guest spaces.', 'Base Cost': 359}},
    {'sheet': '12 - Beds', 'data': {'Supplier Part Number': 'CASA-BED-025', 'Brand': 'Casa Design Furniture', 'Product Name': '#025 JESSICA QUEEN BED', 'Marketing Copy': 'Elegant Jessica series queen bed with sophisticated design details.', 'Base Cost': 1499}},
    {'sheet': '12 - Beds', 'data': {'Supplier Part Number': 'CASA-BED-004', 'Brand': 'Casa Design Furniture', 'Product Name': '#004 SWEETHEART TEEN LOFT BED', 'Marketing Copy': 'Fun and functional loft bed designed specifically for teen bedrooms.', 'Base Cost': 499}},
    # DRESSERS & CHESTS (Category 13)
    {'sheet': '13 - Dressers & Chests', 'data': {'Supplier Part Number': 'CASA-DRS-101', 'Brand': 'Casa Design Furniture', 'Product Name': 'MODERN 6-DRAWER DRESSER', 'Marketing Copy': 'Sleek 6-drawer dresser with smooth-gliding tracks and modern hardware.', 'Base Cost': 550}},
    {'sheet': '13 - Dressers & Chests', 'data': {'Supplier Part Number': 'CASA-DRS-102', 'Brand': 'Casa Design Furniture', 'Product Name': 'ELITE MIRRORED CHEST', 'Marketing Copy': 'High-end mirrored chest that adds light and luxury to any bedroom.', 'Base Cost': 650}},
]

def fill_row(ws, data):
    # Headers are in row 4
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=4, column=col).value
        if h: headers[h] = col
    
    # Find next empty row starting from 6
    row = 6
    while ws.cell(row=row, column=headers.get('Supplier Part Number', 2)).value:
        row += 1
        
    for k, v in data.items():
        if k in headers:
            ws.cell(row=row, column=headers[k], value=v)

for p in products:
    if p['sheet'] in wb.sheetnames:
        fill_row(wb[p['sheet']], p['data'])

wb.save(output_path)
